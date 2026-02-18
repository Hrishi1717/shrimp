from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, field_validator
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import qrcode
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============ Pydantic Models ============

class UserCreate(BaseModel):
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "staff"  # farmer, staff, admin

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str
    created_at: datetime

class SessionCreate(BaseModel):
    session_id: str

class SessionResponse(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str
    session_token: str

class FarmerCreate(BaseModel):
    name: str
    contact: str
    address: str

class Farmer(BaseModel):
    farmer_id: str
    user_id: Optional[str] = None
    name: str
    contact: str
    address: str
    created_at: datetime

class BatchCreate(BaseModel):
    farmer_id: str
    weight_kg: float
    size_grade: str
    location: str

    @field_validator('weight_kg')
    @classmethod
    def validate_weight(cls, v):
        if v <= 0:
            raise ValueError('Weight must be positive')
        return v

class Batch(BaseModel):
    batch_id: str
    farmer_id: str
    weight_kg: float
    size_grade: str
    intake_date: datetime
    intake_time: str
    location: str
    status: str
    qr_code: str
    created_at: datetime

class ProcessingStageCreate(BaseModel):
    batch_id: str
    stage_name: str
    assigned_person: str
    input_weight: float
    output_weight: float

class ProcessingStage(BaseModel):
    stage_id: str
    batch_id: str
    stage_name: str
    assigned_person: str
    input_weight: float
    output_weight: float
    wastage: float
    status: str
    created_at: datetime
    completed_at: Optional[datetime] = None

class InventoryCreate(BaseModel):
    batch_id: str
    location: str
    quantity: float

class Inventory(BaseModel):
    inventory_id: str
    batch_id: str
    location: str
    quantity: float
    batch_age: int
    status: str
    created_at: datetime

class DispatchCreate(BaseModel):
    batch_id: str
    customer_name: str
    country: str
    selling_price: float
    dispatch_date: datetime

class Dispatch(BaseModel):
    dispatch_id: str
    batch_id: str
    customer_name: str
    country: str
    selling_price: float
    dispatch_date: datetime
    status: str
    created_at: datetime

class PaymentCreate(BaseModel):
    farmer_id: str
    batch_id: str
    price_per_kg: float
    deductions: float = 0

class Payment(BaseModel):
    payment_id: str
    farmer_id: str
    batch_id: str
    total_prawns: float
    price_per_kg: float
    gross_amount: float
    deductions: float
    net_amount: float
    payment_status: str
    payment_date: Optional[datetime] = None
    created_at: datetime

# ============ Helper Functions ============

def generate_qr_code(data: dict) -> str:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(json.dumps(data))
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    import base64
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

async def get_current_user(request: Request) -> dict:
    # Check session_token from cookie first, then Authorization header
    session_token = request.cookies.get("session_token")
    
    if not session_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Find session in database
    session_doc = await db.user_sessions.find_one(
        {"session_token": session_token},
        {"_id": 0}
    )
    
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    # Get user
    user_doc = await db.users.find_one(
        {"user_id": session_doc["user_id"]},
        {"_id": 0}
    )
    
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user_doc

# ============ Auth Routes ============

@api_router.post("/auth/session", response_model=SessionResponse)
async def create_session(session_data: SessionCreate, response: Response):
    # REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH
    # Call Emergent Auth API
    import requests
    
    auth_response = requests.get(
        "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
        headers={"X-Session-ID": session_data.session_id}
    )
    
    if auth_response.status_code != 200:
        raise HTTPException(status_code=401, detail="Invalid session ID")
    
    auth_data = auth_response.json()
    
    # Check if user exists
    existing_user = await db.users.find_one(
        {"email": auth_data["email"]},
        {"_id": 0}
    )
    
    if existing_user:
        user_id = existing_user["user_id"]
        # Update user data
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "name": auth_data["name"],
                "picture": auth_data.get("picture")
            }}
        )
    else:
        # Create new user with default role
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": auth_data["email"],
            "name": auth_data["name"],
            "picture": auth_data.get("picture"),
            "role": "staff",  # Default role
            "created_at": datetime.now(timezone.utc)
        })
    
    # Create session
    session_token = auth_data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    })
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    
    return SessionResponse(
        user_id=user_doc["user_id"],
        email=user_doc["email"],
        name=user_doc["name"],
        picture=user_doc.get("picture"),
        role=user_doc["role"],
        session_token=session_token
    )

@api_router.get("/auth/me", response_model=User)
async def get_me(user: dict = Depends(get_current_user)):
    return User(**user)

@api_router.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    # Delete session from database
    await db.user_sessions.delete_many({"user_id": user["user_id"]})
    
    # Clear cookie
    response.delete_cookie(key="session_token", path="/")
    
    return {"message": "Logged out successfully"}

# ============ User Management Routes ============

@api_router.post("/users/invite")
async def invite_user(data: dict, user: dict = Depends(get_current_user)):
    if user["role"] not in ["owner", "admin"]:
        raise HTTPException(status_code=403, detail="Owner/Admin access required")
    
    email = data.get("email")
    role = data.get("role", "staff")
    
    if not email:
        raise HTTPException(status_code=400, detail="Email required")
    
    if role not in ["owner", "admin", "staff", "farmer"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    if existing_user:
        # Update role if user exists
        await db.users.update_one(
            {"email": email},
            {"$set": {"role": role}}
        )
        return {"message": "User role updated", "user_id": existing_user["user_id"]}
    
    # Create invited user record
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    await db.users.insert_one({
        "user_id": user_id,
        "email": email,
        "name": email.split("@")[0],  # Temporary name
        "picture": None,
        "role": role,
        "invited": True,
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"message": "User invited successfully", "user_id": user_id}

@api_router.get("/users", response_model=List[User])
async def get_users(user: dict = Depends(get_current_user)):
    if user["role"] not in ["owner", "admin"]:
        raise HTTPException(status_code=403, detail="Owner/Admin access required")
    
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    
    return users

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: str, user: dict = Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if role not in ["admin", "staff", "farmer"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": role}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Role updated successfully"}

@api_router.post("/farmers/link")
async def link_farmer_to_user(
    data: dict,
    user: dict = Depends(get_current_user)
):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user_id = data.get("user_id")
    farmer_id = data.get("farmer_id")
    
    if not user_id or not farmer_id:
        raise HTTPException(status_code=400, detail="user_id and farmer_id required")
    
    # Update farmer with user_id
    farmer_result = await db.farmers.update_one(
        {"farmer_id": farmer_id},
        {"$set": {"user_id": user_id}}
    )
    
    if farmer_result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    # Update user role to farmer
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": "farmer"}}
    )
    
    return {"message": "Farmer linked to user successfully"}

# ============ Farmer Routes ============

@api_router.post("/farmers", response_model=Farmer)
async def create_farmer(farmer: FarmerCreate, user: dict = Depends(get_current_user)):
    farmer_id = f"farmer_{uuid.uuid4().hex[:12]}"
    farmer_doc = {
        "farmer_id": farmer_id,
        "user_id": None,
        "name": farmer.name,
        "contact": farmer.contact,
        "address": farmer.address,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.farmers.insert_one(farmer_doc)
    
    return Farmer(**farmer_doc)

@api_router.get("/farmers", response_model=List[Farmer])
async def get_farmers(user: dict = Depends(get_current_user)):
    farmers = await db.farmers.find({}, {"_id": 0}).to_list(1000)
    
    for farmer in farmers:
        if isinstance(farmer.get('created_at'), str):
            farmer['created_at'] = datetime.fromisoformat(farmer['created_at'])
    
    return farmers

@api_router.get("/farmers/me/stats")
async def get_farmer_stats(user: dict = Depends(get_current_user)):
    if user["role"] != "farmer":
        raise HTTPException(status_code=403, detail="Not a farmer")
    
    # Find farmer record linked to this user
    farmer = await db.farmers.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not farmer:
        return {
            "total_batches": 0,
            "total_prawns_supplied": 0,
            "total_payments": 0,
            "pending_payments": 0
        }
    
    # Get batches for this farmer
    batches = await db.batches.find({"farmer_id": farmer["farmer_id"]}, {"_id": 0}).to_list(1000)
    total_prawns = sum(b["weight_kg"] for b in batches)
    
    # Get payments
    payments = await db.payments.find({"farmer_id": farmer["farmer_id"]}, {"_id": 0}).to_list(1000)
    total_paid = sum(p["net_amount"] for p in payments if p["payment_status"] == "paid")
    pending_payments = sum(p["net_amount"] for p in payments if p["payment_status"] == "pending")
    
    return {
        "total_batches": len(batches),
        "total_prawns_supplied": total_prawns,
        "total_payments": total_paid,
        "pending_payments": pending_payments,
        "payments": payments
    }

# ============ Batch Routes ============

@api_router.post("/batches", response_model=Batch)
async def create_batch(batch: BatchCreate, user: dict = Depends(get_current_user)):
    batch_id = f"BATCH{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"
    
    # Generate QR code
    qr_data = {
        "batch_id": batch_id,
        "farmer_id": batch.farmer_id,
        "weight_kg": batch.weight_kg,
        "size_grade": batch.size_grade,
        "intake_date": datetime.now(timezone.utc).isoformat()
    }
    qr_code = generate_qr_code(qr_data)
    
    batch_doc = {
        "batch_id": batch_id,
        "farmer_id": batch.farmer_id,
        "weight_kg": batch.weight_kg,
        "size_grade": batch.size_grade,
        "intake_date": datetime.now(timezone.utc),
        "intake_time": datetime.now(timezone.utc).strftime("%H:%M:%S"),
        "location": batch.location,
        "status": "RECEIVED",
        "qr_code": qr_code,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.batches.insert_one(batch_doc)
    
    return Batch(**batch_doc)

@api_router.get("/batches", response_model=List[Batch])
async def get_batches(user: dict = Depends(get_current_user)):
    batches = await db.batches.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for batch in batches:
        if isinstance(batch.get('intake_date'), str):
            batch['intake_date'] = datetime.fromisoformat(batch['intake_date'])
        if isinstance(batch.get('created_at'), str):
            batch['created_at'] = datetime.fromisoformat(batch['created_at'])
    
    return batches

@api_router.get("/batches/{batch_id}", response_model=Batch)
async def get_batch(batch_id: str, user: dict = Depends(get_current_user)):
    batch = await db.batches.find_one({"batch_id": batch_id}, {"_id": 0})
    
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    if isinstance(batch.get('intake_date'), str):
        batch['intake_date'] = datetime.fromisoformat(batch['intake_date'])
    if isinstance(batch.get('created_at'), str):
        batch['created_at'] = datetime.fromisoformat(batch['created_at'])
    
    return Batch(**batch)

# ============ Processing Routes ============

@api_router.post("/processing", response_model=ProcessingStage)
async def create_processing_stage(stage: ProcessingStageCreate, user: dict = Depends(get_current_user)):
    stage_id = f"stage_{uuid.uuid4().hex[:12]}"
    
    wastage = stage.input_weight - stage.output_weight
    yield_percentage = (stage.output_weight / stage.input_weight * 100) if stage.input_weight > 0 else 0
    
    stage_doc = {
        "stage_id": stage_id,
        "batch_id": stage.batch_id,
        "stage_name": stage.stage_name,
        "assigned_person": stage.assigned_person,
        "input_weight": stage.input_weight,
        "output_weight": stage.output_weight,
        "wastage": wastage,
        "yield_percentage": yield_percentage,
        "status": "COMPLETED",
        "created_at": datetime.now(timezone.utc),
        "completed_at": datetime.now(timezone.utc)
    }
    
    await db.processing_stages.insert_one(stage_doc)
    
    # Update batch status
    stages_count = await db.processing_stages.count_documents({"batch_id": stage.batch_id})
    if stages_count >= 4:  # All 4 stages completed
        await db.batches.update_one(
            {"batch_id": stage.batch_id},
            {"$set": {"status": "PROCESSED"}}
        )
    
    return ProcessingStage(**stage_doc)

@api_router.get("/processing/batch/{batch_id}", response_model=List[ProcessingStage])
async def get_processing_stages(batch_id: str, user: dict = Depends(get_current_user)):
    stages = await db.processing_stages.find({"batch_id": batch_id}, {"_id": 0}).to_list(1000)
    
    for stage in stages:
        if isinstance(stage.get('created_at'), str):
            stage['created_at'] = datetime.fromisoformat(stage['created_at'])
        if stage.get('completed_at') and isinstance(stage.get('completed_at'), str):
            stage['completed_at'] = datetime.fromisoformat(stage['completed_at'])
    
    return stages

# ============ Inventory Routes ============

@api_router.post("/inventory", response_model=Inventory)
async def create_inventory(inventory: InventoryCreate, user: dict = Depends(get_current_user)):
    inventory_id = f"inv_{uuid.uuid4().hex[:12]}"
    
    # Get batch to calculate age
    batch = await db.batches.find_one({"batch_id": inventory.batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    intake_date = batch["intake_date"]
    if isinstance(intake_date, str):
        intake_date = datetime.fromisoformat(intake_date)
    
    batch_age = (datetime.now(timezone.utc) - intake_date).days
    
    inventory_doc = {
        "inventory_id": inventory_id,
        "batch_id": inventory.batch_id,
        "location": inventory.location,
        "quantity": inventory.quantity,
        "batch_age": batch_age,
        "status": "STORED",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.inventory.insert_one(inventory_doc)
    
    # Update batch status
    await db.batches.update_one(
        {"batch_id": inventory.batch_id},
        {"$set": {"status": "STORED"}}
    )
    
    return Inventory(**inventory_doc)

@api_router.get("/inventory", response_model=List[Inventory])
async def get_inventory(user: dict = Depends(get_current_user)):
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    
    for item in inventory:
        if isinstance(item.get('created_at'), str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
    
    return inventory

# ============ Dispatch Routes ============

@api_router.post("/dispatch", response_model=Dispatch)
async def create_dispatch(dispatch: DispatchCreate, user: dict = Depends(get_current_user)):
    dispatch_id = f"disp_{uuid.uuid4().hex[:12]}"
    
    dispatch_doc = {
        "dispatch_id": dispatch_id,
        "batch_id": dispatch.batch_id,
        "customer_name": dispatch.customer_name,
        "country": dispatch.country,
        "selling_price": dispatch.selling_price,
        "dispatch_date": dispatch.dispatch_date,
        "status": "SHIPPED",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.dispatches.insert_one(dispatch_doc)
    
    # Update batch status
    await db.batches.update_one(
        {"batch_id": dispatch.batch_id},
        {"$set": {"status": "SHIPPED"}}
    )
    
    return Dispatch(**dispatch_doc)

@api_router.get("/dispatch", response_model=List[Dispatch])
async def get_dispatches(user: dict = Depends(get_current_user)):
    dispatches = await db.dispatches.find({}, {"_id": 0}).to_list(1000)
    
    for dispatch in dispatches:
        if isinstance(dispatch.get('dispatch_date'), str):
            dispatch['dispatch_date'] = datetime.fromisoformat(dispatch['dispatch_date'])
        if isinstance(dispatch.get('created_at'), str):
            dispatch['created_at'] = datetime.fromisoformat(dispatch['created_at'])
    
    return dispatches

# ============ Payment Routes ============

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment: PaymentCreate, user: dict = Depends(get_current_user)):
    payment_id = f"pay_{uuid.uuid4().hex[:12]}"
    
    # Get batch to calculate total
    batch = await db.batches.find_one({"batch_id": payment.batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    total_prawns = batch["weight_kg"]
    gross_amount = total_prawns * payment.price_per_kg
    net_amount = gross_amount - payment.deductions
    
    payment_doc = {
        "payment_id": payment_id,
        "farmer_id": payment.farmer_id,
        "batch_id": payment.batch_id,
        "total_prawns": total_prawns,
        "price_per_kg": payment.price_per_kg,
        "gross_amount": gross_amount,
        "deductions": payment.deductions,
        "net_amount": net_amount,
        "payment_status": "pending",
        "payment_date": None,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.payments.insert_one(payment_doc)
    
    return Payment(**payment_doc)

@api_router.put("/payments/{payment_id}/status")
async def update_payment_status(payment_id: str, status: str, user: dict = Depends(get_current_user)):
    result = await db.payments.update_one(
        {"payment_id": payment_id},
        {"$set": {
            "payment_status": status,
            "payment_date": datetime.now(timezone.utc) if status == "paid" else None
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment status updated"}

@api_router.get("/payments", response_model=List[Payment])
async def get_payments(user: dict = Depends(get_current_user)):
    payments = await db.payments.find({}, {"_id": 0}).to_list(1000)
    
    for payment in payments:
        if isinstance(payment.get('created_at'), str):
            payment['created_at'] = datetime.fromisoformat(payment['created_at'])
        if payment.get('payment_date') and isinstance(payment.get('payment_date'), str):
            payment['payment_date'] = datetime.fromisoformat(payment['payment_date'])
    
    return payments

# ============ Dashboard Routes ============

@api_router.get("/dashboard/admin")
async def get_admin_dashboard(user: dict = Depends(get_current_user)):
    # Total procurement
    batches = await db.batches.find({}, {"_id": 0}).to_list(10000)
    total_procurement = sum(b["weight_kg"] for b in batches)
    
    # Calculate yield
    stages = await db.processing_stages.find({}, {"_id": 0}).to_list(10000)
    total_input = sum(s["input_weight"] for s in stages)
    total_output = sum(s["output_weight"] for s in stages)
    yield_percentage = (total_output / total_input * 100) if total_input > 0 else 0
    
    # Farmer payments
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    total_payments = sum(p["net_amount"] for p in payments)
    pending_payments = sum(p["net_amount"] for p in payments if p["payment_status"] == "pending")
    
    # Selling price (average from dispatches)
    dispatches = await db.dispatches.find({}, {"_id": 0}).to_list(10000)
    avg_selling_price = sum(d["selling_price"] for d in dispatches) / len(dispatches) if dispatches else 0
    
    return {
        "total_procurement": total_procurement,
        "yield_percentage": yield_percentage,
        "total_payments": total_payments,
        "pending_payments": pending_payments,
        "avg_selling_price": avg_selling_price,
        "total_batches": len(batches),
        "total_farmers": await db.farmers.count_documents({}),
        "total_dispatches": len(dispatches)
    }

# ============ Export Routes ============

def create_styled_header(worksheet, headers: List[str]):
    header_fill = PatternFill(
        start_color="0F172A",
        end_color="0F172A",
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        worksheet.column_dimensions[cell.column_letter].width = 18
    
    return worksheet

@api_router.post("/export/batches")
async def export_batches(user: dict = Depends(get_current_user)):
    batches = await db.batches.find({}, {"_id": 0}).to_list(10000)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Batches"
    
    headers = ["Batch ID", "Farmer ID", "Weight (kg)", "Size Grade", "Intake Date", "Location", "Status"]
    worksheet = create_styled_header(worksheet, headers)
    
    for row_num, batch in enumerate(batches, 2):
        intake_date = batch["intake_date"]
        if isinstance(intake_date, str):
            intake_date = datetime.fromisoformat(intake_date)
        
        row_data = [
            batch["batch_id"],
            batch["farmer_id"],
            batch["weight_kg"],
            batch["size_grade"],
            intake_date.strftime("%Y-%m-%d %H:%M:%S"),
            batch["location"],
            batch["status"]
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    filename = f"batches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/export/payments")
async def export_payments(user: dict = Depends(get_current_user)):
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payments"
    
    headers = ["Payment ID", "Farmer ID", "Batch ID", "Total Prawns (kg)", "Price/kg", "Gross Amount", "Deductions", "Net Amount", "Status"]
    worksheet = create_styled_header(worksheet, headers)
    
    for row_num, payment in enumerate(payments, 2):
        row_data = [
            payment["payment_id"],
            payment["farmer_id"],
            payment["batch_id"],
            payment["total_prawns"],
            payment["price_per_kg"],
            payment["gross_amount"],
            payment["deductions"],
            payment["net_amount"],
            payment["payment_status"]
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if col_num in [5, 6, 7, 8]:  # Currency columns
                cell.number_format = '$#,##0.00'
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    filename = f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/export/processing")
async def export_processing(user: dict = Depends(get_current_user)):
    stages = await db.processing_stages.find({}, {"_id": 0}).to_list(10000)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Processing Stages"
    
    headers = ["Stage ID", "Batch ID", "Stage Name", "Assigned Person", "Input Weight", "Output Weight", "Wastage", "Yield %", "Status"]
    worksheet = create_styled_header(worksheet, headers)
    
    for row_num, stage in enumerate(stages, 2):
        yield_pct = stage.get("yield_percentage", 0)
        
        row_data = [
            stage["stage_id"],
            stage["batch_id"],
            stage["stage_name"],
            stage["assigned_person"],
            stage["input_weight"],
            stage["output_weight"],
            stage["wastage"],
            f"{yield_pct:.2f}%",
            stage["status"]
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    filename = f"processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
